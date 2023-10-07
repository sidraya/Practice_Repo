import openpyxl
import requests

file1="S:\GetAPI.xlsx"
workbook=openpyxl.load_workbook(file1)
sheet=workbook.active
def get_row_count(file1,sheet):
    return sheet.max_row
def get_column_count(file1,sheet):
    return sheet.max_column
r=get_row_count(file1,sheet)
c=get_column_count(file1,sheet)

for i in range(2,r+1):
    for j in range (2,c+1):
        url=sheet.cell(2,1).value
        resp=requests.get(url)
        time=resp.elapsed.total_seconds()
        sheet.cell(i ,j).value=time
        workbook.save(file1)