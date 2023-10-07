import openpyxl
import requests

file1='S:\GetPutPatch.xlsx'
workbook=openpyxl.load_workbook(file1)
sheet=workbook.active
def get_row_count(file1,sheet):
    return sheet.max_row
def get_column_count(file1,sheet):
    return sheet.max_column
r=get_row_count(file1,sheet)
c=get_column_count(file1,sheet)

for i in range(2,r+1):
    url=sheet.cell(2,1).value
    payload=sheet.cell(2,2).value
    url1=sheet.cell(5,1).value
    payload1=sheet.cell(5,2).value
    url2=sheet.cell(6,1).value
    payload2=sheet.cell(6,2).value

    print("Url of post method=",url)
    print("Payload of post method=",payload)
    resp=requests.post(url,payload)
    resp1=requests.put(url1,payload1)
    resp2=requests.patch(url2,payload2)
    code=resp.status_code
    time=resp.elapsed.total_seconds()
    code1=resp1.status_code
    time1=resp1.elapsed.total_seconds()
    code2=resp2.status_code
    time2=resp2.elapsed.total_seconds()

    print("Status code of post method=",code)
    sheet.cell(i,3).value=code
    sheet.cell(i,4).value=time
    sheet.cell(5,3).value=code1
    sheet.cell(5,4).value=time1
    sheet.cell(6,3).value=code2
    sheet.cell(6,4).value=time2
    workbook.save(file1)
    print(resp.status_code)


